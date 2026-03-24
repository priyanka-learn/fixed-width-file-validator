"""
Fixed Width File (FWF) Validator
Validates fixed-width export files against an Excel-based field definition schema.
Supports Customer/Party, Account, Transaction, Relationship export files
commonly used in banking and insurance domain.
"""

import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

excfilename = input("Input Excel workbook name with extension: ")
filespath = input("Input file's path: ")

expfilelist = os.listdir(filespath)
filecount = 0
files = []

for fl in expfilelist:
    if fl[-4:].upper() == ".TXT":
        files.append(fl)
        filecount = 1 + filecount
        expfilename = fl
        print("Evaluating file - ", expfilename)
        errorfile = 0

        Data = pd.read_excel(filespath + "\\" + excfilename, sheet_name='Sheet1', header=0, index_col=False)

        workbook = load_workbook(filename=filespath + "\\" + excfilename)
        worksheet: Worksheet = workbook.active

        """Create Folder - Validated Files"""
        if not os.path.exists(filespath + '\\Validated_Files'):
            os.makedirs(filespath + '\\Validated_Files')

        """Create Result File in 'Validated Files' folder"""
        # FIX: Using replace instead of strip to avoid character-level stripping
        expfilenamestrip = expfilename.replace('.txt', '').replace('.TXT', '')
        resultfile = open(filespath + "\\Validated_Files" + "\\" + expfilenamestrip + "_Validated.txt", encoding="utf-16", mode="w+")
        filelogno = 1

        """Getting total number of columns from source schema excel"""
        ColCount = Data.iloc[:, 1].count()
        resultfile.write(str(filelogno) + ". No of columns as per excel sheet: " + str(ColCount) + "\n")
        filelogno = filelogno + 1

        """Validate record lengths of Fixed width file"""
        RecLen = Data.iloc[:, 3].sum()

        try:  # to avoid UTF error
            with open(filespath + "\\" + expfilename, encoding='utf-8', mode='r+') as f:
                lines = f.readlines()

        except ValueError:
            with open(filespath + "\\" + expfilename, encoding='utf-16', mode='r+') as f:
                lines = f.readlines()

        resultfile.write(str(filelogno) + ". Expected Record length as per excel sheet: " + str(RecLen) + "\n")
        filelogno = filelogno + 1
        RecLenFlag = 0

        for i in lines:
            if len(i[1:]) != RecLen:
                line_no = lines.index(i)
                RecLenFlag = 1
                resultfile.write(str(filelogno) + ". Record length incorrect as:" + str(len(i[1:])) + " in Row Number:" + str(line_no + 1) + "\n")
                filelogno = filelogno + 1
                errorfile = 1

        if RecLenFlag == 0:
            resultfile.write(str(filelogno) + ". All record's length is correct as: " + str(RecLen) + "\n")
            filelogno = filelogno + 1

        resultfile.write("\n")

        """Calculating Slicing values from source schema excel as per column width"""
        # FIX: Pure Python calculation — no Excel formulas, no COM automation needed
        start = 0
        for r in range(ColCount):
            field_len = worksheet['$D' + str(r + 2)].value
            worksheet['$F' + str(r + 2)] = start
            worksheet['$G' + str(r + 2)] = start + field_len
            start += field_len
        workbook.save(filename=filespath + "\\" + excfilename)

        """ Checking if TRXN file """
        if "TRXN" in expfilename.upper():
            flines = lines[1:]
        else:
            flines = lines

        """Traversing Each record"""
        workbook = load_workbook(filename=filespath + "\\" + excfilename)
        worksheet: Worksheet = workbook.active

        RecCount = 0

        # FIX: Using enumerate for correct line numbers throughout
        for n, i in enumerate(flines):
            # FIX: Calculate actual line number correctly
            # For TRXN files, flines starts from index 1 of original lines, so offset by 1
            if "TRXN" in expfilename.upper():
                line_no = n + 1  # +1 because we skipped header row
            else:
                line_no = n

            RecCount = RecCount + 1
            ary = []

            for r in range(ColCount):
                sta = worksheet['$F' + str(r + 2)].value
                sto = worksheet['$G' + str(r + 2)].value
                datavalue = i[slice(sta, sto)]
                ary.append(i[slice(sta, sto)])

                # Read datatype once for cleaner comparisons
                datatype = worksheet['$C' + str(r + 2)].value.upper().strip()
                mandatory = worksheet['$E' + str(r + 2)].value.upper().strip()
                field_name = worksheet['$B' + str(r + 2)].value.upper()
                field_length = worksheet['$D' + str(r + 2)].value

                """Date Validation"""
                # Date fields always contain YYYYMMDDHHmmSS (14 chars), right-aligned within field length

                """Mandatory Date/Datetime validation"""
                if datatype in ('DATE', 'DATETIME') and mandatory == 'Y':

                    dt_string = datavalue

                    if len(dt_string.strip()) == 14:
                        try:
                            dt_object = datetime.strptime(dt_string.strip(), "%Y%m%d%H%M%S")

                        except ValueError:
                            resultfile.write(' '.join([str(filelogno) + ".", field_name,
                                "- Invalid mandatory Date string as:", str(dt_string), "in Row:",
                                str(line_no + 1) + ",", "Column:", str(r + 1) + ",", "Start position:", str(sta + 1),
                                "and Size:", str(field_length) + "\n"]))
                            filelogno = filelogno + 1
                            errorfile = 1

                    else:
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Invalid mandatory Date string as:", str(dt_string), "in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                """Non-Mandatory Date/Datetime validation"""
                if datatype in ('DATE', 'DATETIME') and mandatory == 'N' and datavalue.strip() != '':
                    dt_string = datavalue
                    if len(dt_string.strip()) == 14:
                        try:
                            dt_object = datetime.strptime(dt_string.strip(), "%Y%m%d%H%M%S")
                        except ValueError:
                            resultfile.write(' '.join([str(filelogno) + ".", field_name,
                                "- Invalid Date string as:", str(dt_string), "in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                                "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                            filelogno = filelogno + 1
                            errorfile = 1
                    elif len(dt_string.strip()) != 0:
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Invalid Date string as:", str(dt_string), "in Row:",
                            str(line_no + 1) + ",", "Column:", str(r + 1) + ",", "Start position:", str(sta + 1),
                            "and Size:", str(field_length) + "\n"]))
                        # FIX: These were outside the elif block before, causing false errors
                        filelogno = filelogno + 1
                        errorfile = 1

                """Date/Datetime Whitespace padding validation"""
                if datatype in ('DATE', 'DATETIME') and datavalue.strip() != '':
                    ExpSpc = field_length - 14
                    dt_string = datavalue
                    Spacespan = re.search(r'[\s]+', str(dt_string))
                    if ExpSpc != 0 and Spacespan is not None and Spacespan.span() != (0, ExpSpc):
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- White Spaces not as expected in:", str(dt_string), "in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                """AMOUNT Validation"""
                if datatype == 'AMOUNT' and datavalue.strip() != '':
                    amt_string = datavalue.strip()
                    try:
                        if len(amt_string) < 3:
                            resultfile.write(' '.join([str(filelogno) + ".", field_name,
                                "- Amount value less than three digit or blank in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                                "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                            filelogno = filelogno + 1
                            errorfile = 1
                        elif len(amt_string) >= 3:
                            amt_object = int(amt_string)
                    except ValueError:
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Amount not an integer in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                if datatype == 'AMOUNT' and datavalue.strip() != '':
                    amt_str = datavalue
                    amt_string = datavalue.strip()
                    ExpSpc = field_length - len(amt_string)
                    Spacespan = re.search(r'[\s]+', str(amt_str))
                    if ExpSpc != 0 and Spacespan is not None and Spacespan.span() != (0, ExpSpc):
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- White Spaces not as expected in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                elif datatype == 'AMOUNT' and mandatory == 'Y' and datavalue.strip() == '':
                    resultfile.write(' '.join([str(filelogno) + ".", field_name,
                        "- Mandatory Amount field blank in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                        "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                    filelogno = filelogno + 1
                    errorfile = 1

                """TEXT Validation"""
                if datatype == 'TEXT' and "NOT" not in field_name and "USED" not in field_name:
                    txt_string = datavalue
                    Spacespan = re.search(r'[\s]+', str(txt_string))
                    if txt_string.strip() != '' and Spacespan is not None and Spacespan.start() == 0:
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- White Spaces not as expected in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",", "Start position:",
                            str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                    elif txt_string.strip() == '' and mandatory == 'Y':
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Mandatory value blank in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",", "Start position:",
                            str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                if datatype == 'TEXT' and "NOT" in field_name and "USED" in field_name:
                    if datavalue.strip() != '':
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Not Used column contains value in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1

                """NUMBER Validation"""
                if datatype in ('NUMERIC', 'NUMBER', 'INT'):
                    num_string = datavalue
                    if num_string.strip() != '':
                        try:
                            num_object = int(num_string)
                        except ValueError:
                            resultfile.write(' '.join([str(filelogno) + ".", field_name,
                                "- Value not as Integer in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                                "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                            filelogno = filelogno + 1
                            errorfile = 1
                    elif num_string.strip() == '' and mandatory == 'Y':
                        resultfile.write(' '.join([str(filelogno) + ".", field_name,
                            "- Mandatory value blank in Row:", str(line_no + 1) + ",", "Column:", str(r + 1) + ",",
                            "Start position:", str(sta + 1), "and Size:", str(field_length) + "\n"]))
                        filelogno = filelogno + 1
                        errorfile = 1
            workbook.save(excfilename)
            resultfile.write("\n")

        """Validating TRXN file's Header"""
        # Note- header format: 9 zeros + 'HEADER' + 5 digit run number + 10 digit record count
        # e.g- 000000000HEADER000010000001000
        if "TRXN" in expfilename.upper():
            header = lines[0]
            runnum = expfilename[-9:]

            if header[0:9] == '000000000':
                resultfile.write(str(filelogno) + ". " + "HEADER- Initial nine zeros as expected: " + header[0:9] + "\n")
                filelogno = filelogno + 1
            else:
                resultfile.write(str(filelogno) + ". " + "HEADER- Nine zeros NOT as expected: " + header[0:9] + "\n")
                filelogno = filelogno + 1
                errorfile = 1

            if header[9:15] == "HEADER":
                resultfile.write(str(filelogno) + ". " + "HEADER- Header as  expected: " + header[9:15] + "\n")
                filelogno = filelogno + 1
            else:
                resultfile.write(str(filelogno) + ". " + "HEADER- Header NOT as expected: " + header[9:15] + "\n")
                filelogno = filelogno + 1
                errorfile = 1

            if header[15:20] == runnum[0:5]:
                resultfile.write(str(filelogno) + ". " + "HEADER- Run number as expected: " + header[15:20] + "\n")
                filelogno = filelogno + 1
            else:
                resultfile.write(str(filelogno) + ". " + "HEADER- Run number NOT as expected: " + header[15:20] + "\n")
                filelogno = filelogno + 1
                errorfile = 1

            if header[20:30] == str(RecCount).zfill(10):
                resultfile.write(str(filelogno) + ". " + "HEADER- Record count as expected: " + header[20:30] + "\n")
                filelogno = filelogno + 1
            else:
                resultfile.write(str(filelogno) + ". " + "HEADER- Record count NOT as expected: " + header[20:30] + "\n")
                filelogno = filelogno + 1
                errorfile = 1

        resultfile.write(str(filelogno) + ". " + "Record count in Export File is: " + str(RecCount) + "\n")
        filelogno = filelogno + 1

        resultfile.close()

        if errorfile == 1:
            if not os.path.exists(filespath + '\\Validated_Files\\Error_Files'):
                os.makedirs(filespath + '\\Validated_Files' + '\\Error_Files')
            try:
                os.rename(filespath + "\\Validated_Files" + "\\" + expfilenamestrip + "_Validated.txt",
                          filespath + "\\Validated_Files\\Error_Files" + "\\" + expfilenamestrip + "_Error.txt")
            except WindowsError:
                os.remove(filespath + "\\Validated_Files\\Error_Files" + "\\" + expfilenamestrip + "_Error.txt")
                os.rename(filespath + "\\Validated_Files" + "\\" + expfilenamestrip + "_Validated.txt",
                          filespath + "\\Validated_Files\\Error_Files" + "\\" + expfilenamestrip + "_Error.txt")
